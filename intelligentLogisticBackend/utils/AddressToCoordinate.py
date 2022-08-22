import requests
import json
import geocoder
from openpyxl import load_workbook


class AddressToCoordinate(object):
    '''
    给定地址，获取经纬度
    '''

    def __init__(self, channel):
        self.channel = channel
        self.key = None

    def _gaudurMap(self, address):
        '''
        高德地图
        '''
        url = 'https://restapi.amap.com/v3/geocode/geo'
        # 参数放入字典
        params = {'key': self.key,
                  'address': address}
        res = requests.get(url, params)
        json_text = json.loads(res.text)
        return list(map(float, json_text['geocodes'][0]['location'].split(',')))

    def _baiduMap(self, address):
        '''
        百度地图
        '''
        url = 'http://api.map.baidu.com/geocoding/v3/'
        # 参数放入字典
        params = {'ak': self.key,
                  'address': address,
                  'output': 'json'}
        res = requests.get(url, params)
        json_text = json.loads(res.text)

        return [json_text['result']['location']['lng'], json_text['result']['location']['lat']]

    def _tencentMap(self, address):
        '''
        腾讯地图
        '''
        url = 'https://apis.map.qq.com/jsapi?'
        # 参数放入字典
        params = {
            "qt": "geoc",
            "addr": address,  # 传入地址参数
            "output": "json",
            "key": self.key,  # 即腾讯地图API的key
            "pf": "jsapi",
            "ref": "jsapi"}
        res = requests.get(url, params)
        json_text = json.loads(res.text)

        return list(map(float, [json_text['detail']['pointx'], json_text['detail']['pointy']]))

    def _freeArcgisMap(self, address):
        '''
        免费的Arcgis，速度很慢
        '''
        try:
            result = geocoder.arcgis(address)
            lng_lat = [result.latlng[1], result.latlng[0]]
        except:
            lng_lat = [0.0, 0.0]

        return lng_lat

    def _chooseMap(self):
        '''
        根据界面选择，使用对应的厂商地图
        '''
        if self.channel == '高德':
            map_provider = self._gaudurMap
        elif self.channel == '百度':
            map_provider = self._baiduMap
        elif self.channel == '腾讯':
            map_provider = self._tencentMap
        elif self.channel == 'Arcgis':
            map_provider = self._freeArcgisMap
        else:
            map_provider = None

        return map_provider

    def getSingleCoordinate(self, address):
        '''
        获取单个地址的经纬度
        '''
        map_provider = self._chooseMap()
        coordinate = map_provider(address)

        return coordinate

    def getBatchCoordinate(self, path, sheet_name="需求点"):
        '''
        批量获取地址的经纬度,并把数据更新到当前文件
        '''
        message_dict = {'status': 'success', 'message': ''}
        if sheet_name=='需求点':
            name_col_dict = {'客户名称': 'A', '省份': 'B', "城市": 'C', "详细地址": 'D', '经度': 'E', "纬度": 'F'}
        elif sheet_name == '指定仓':
            name_col_dict = {'省份': 'A', "城市": 'B', "详细地址": 'C', '经度': 'D', "纬度": 'E'}
        try:
            map_provider = self._chooseMap()
            workbook = load_workbook(path)
            sheet = workbook[sheet_name]
            flag = True
            for i in range(2, sheet.max_row+1):
                if sheet[name_col_dict['省份']+str(i)].value is None:
                    break
                address = sheet[name_col_dict['详细地址'] + str(i)].value
                # print('in Utils', address)
                if address is not None:
                    if sheet[name_col_dict['经度']+str(i)].value and sheet[name_col_dict['纬度']+str(i)].value:
                        # print('jingweidu', sheet[name_col_dict['经度']+str(i)].value)
                        continue

                    concat_address = sheet[name_col_dict['省份'] + str(i)].value + sheet[name_col_dict['城市'] + str(i)].value + address.strip()
                    try:
                        coordinate = map_provider(concat_address)
                        sheet[name_col_dict['经度'] + str(i)].value = coordinate[0]
                        sheet[name_col_dict['纬度'] + str(i)].value = coordinate[1]
                    except:
                        message = '部分地址无法解析'
                        flag = False
                        sheet[name_col_dict['经度'] + str(i)].value = None
                        sheet[name_col_dict['纬度'] + str(i)].value = None
                else:
                    message = '详细地址存在空值！！！'
                    message_dict['status'] = 'fail'
                    message_dict['message'] = message
                    return message_dict
            workbook.save(path)
            if flag == True:
                message_dict['status'] = 'success'
                message_dict['message'] = '经纬度获取完成!\n结果存储在{}'.format(path)
                return message_dict
            else:
                message_dict['status'] = 'fail'
                message_dict['message'] = message
                return message_dict

        except Exception as e:
            message_dict['status'] = 'fail'
            message_dict['message'] = str(e)
            return message_dict
